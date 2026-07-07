#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
官網訂單 → 蝦皮訂單格式 轉換工具

用法:
  python sc_to_shopee.py <來源訂單.xlsx> <蝦皮格式.xlsx> [輸出檔名.xlsx]

若未指定輸出檔名，自動產生「蝦皮格式_<原始檔名>.xlsx」
"""
import openpyxl, sys, os, re, io, shutil
from datetime import datetime

# ──────────────────────────────────────────────
# 共用函式
# ──────────────────────────────────────────────

def clean(val):
    """清理來源數值：去除 ==、引號等前綴"""
    if val is None:
        return ''
    s = str(val).strip()
    if s.startswith('='):
        s = s[1:]
    s = s.strip('"')
    s = s.strip("'")
    return s.strip()


def parse_store(addr):
    """從地址擷取超商門市店號與名稱
    範例: 買家超商取貨:久旺 (231196)  →  231196, 久旺
    """
    m = re.search(r'[：:](\D+?)\s*\((\d+)\)', addr)
    if m:
        return m.group(2), m.group(1)
    return '', ''


def parse_city_area(addr):
    """從地址粗略解析城市與行政區"""
    cities = [
        '台北市', '新北市', '桃園市', '台中市', '台南市', '高雄市',
        '基隆市', '新竹市', '新竹縣', '苗栗縣', '彰化縣', '南投縣',
        '雲林縣', '嘉義市', '嘉義縣', '屏東縣', '宜蘭縣', '花蓮縣',
        '台東縣', '澎湖縣', '金門縣', '連江縣'
    ]
    for city in cities:
        if city in addr:
            rest = addr.split(city, 1)[1]
            m = re.match(r'([^\s]{2,3}[區市鎮鄉])', rest)
            if m:
                return city, m.group(1)
            return city, ''
    return '', ''


def map_payment(pay):
    """正規化付款方式為蝦皮用語"""
    if '信用卡' in pay:
        return '信用卡'
    if '轉帳' in pay or '匯款' in pay:
        return '銀行轉帳'
    if 'line' in pay.lower():
        return 'Line Pay'
    return pay


def map_shipping(ship):
    """正規化配送方式為蝦皮用語"""
    for kw, mapped in [('7-11', '7-11 超商取貨'), ('全家', '全家 超商取貨'),
                       ('自取', '賣家自行配送'), ('宅配', '賣家自行配送'),
                       ('郵局', '郵局'), ('宅急便', '宅急便')]:
        if kw in ship:
            return mapped
    return ship


# ──────────────────────────────────────────────
# Phase 1：清理來源訂單
# ──────────────────────────────────────────────

def clean_orders(ws):
    """
    1. 刪除完全空白列
    2. 子項目補上母訂單訂購資訊
    3. 其餘空白補「—」
    4. 從 Referer 推斷 Campaign Source
    """
    # ── 1. 刪除空白列 ──
    empty_rows = []
    for row in range(ws.max_row, 1, -1):
        if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
            empty_rows.append(row)
    for r in empty_rows:
        ws.delete_rows(r)

    if empty_rows:
        print(f"  -> 已刪除 {len(empty_rows)} 列空白: {list(reversed(empty_rows))}")

    # ── 2. 子項目補母訂單資訊 ──
    fill_fields = {
        1: '訂單編號', 2: '交易編號', 3: '成立日期',
        4: '出貨狀態', 5: '付款狀態', 6: '訂單狀態', 7: '付款方式',
        8: '訂單金額', 9: '應付金額',
        10: '購買人姓名', 11: '購買人電話', 12: '購買人email',
        13: '備註', 14: '賣家備註',
        15: '配送方式', 16: '收件人姓名', 17: '收件人手機', 18: '收件人地址',
        19: '出貨單號', 20: '出貨日期', 21: '配送時段',
        22: '發票形式', 23: '發票抬頭', 24: '統一編號',
        25: '商品總計', 26: '運費', 27: '折扣', 28: '手續費', 29: '額外費用', 30: '合計',
        40: '商品狀態',
        42: '銷售頁代稱', 43: '優惠碼名稱', 44: '優惠碼代稱',
        45: '參照位址Referer', 46: 'Campaign Source', 47: 'Campaign Medium',
        48: 'Campaign Name', 49: 'Campaign ID', 50: 'Term', 51: 'Content',
        52: '美安RID', 53: '美安Click_ID',
        54: '會員購買次數', 55: '同電話購買次數', 56: '同Email購買次數',
        57: '來就省取貨店號', 58: '團訂必填-用書學校',
        59: '團訂必填-上課班級', 60: '團訂必填-授課教師'
    }

    parent = {}
    sub_count = 0
    for row in range(2, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        ae_val = ws.cell(row=row, column=31).value

        if a_val and str(a_val).strip():
            parent = {col: ws.cell(row=row, column=col).value for col in fill_fields}
        elif ae_val and str(ae_val).strip() and parent:
            filled = 0
            for col in fill_fields:
                cur = ws.cell(row=row, column=col).value
                if cur is None or str(cur).strip() == '':
                    pv = parent.get(col)
                    if pv is not None and str(pv).strip():
                        ws.cell(row=row, column=col).value = pv
                        filled += 1
            if filled:
                sub_count += 1
                print(f"  -> Row {row}: 子項「{str(ae_val).strip()[:20]}」-> 補入 {filled} 欄")

    if sub_count:
        print(f"  -> 共補齊 {sub_count} 列子項目訂購資訊")

    # ── 3. 其餘空白補「—」及推斷 ──
    mark_cols = {
        13: '備註', 14: '賣家備註', 19: '出貨單號',
        34: '商品組合', 35: '商品型號', 36: '商品品牌',
        41: '商品備註', 42: '銷售頁代稱',
        43: '優惠碼名稱', 44: '優惠碼代稱',
        45: '參照位址Referer', 46: 'Campaign Source',
        47: 'Campaign Medium', 48: 'Campaign Name', 49: 'Campaign ID',
        50: 'Term', 51: 'Content',
        52: '美安RID', 53: '美安Click_ID',
        57: '來就省取貨店號', 58: '團訂必填-用書學校',
        59: '團訂必填-上課班級', 60: '團訂必填-授課教師'
    }

    src_map = {'google': 'google', 'line': 'line', 'facebook': 'facebook', 'fb': 'facebook',
               'gm': 'email', 'gmail': 'email', 'mail': 'email',
               'instagram': 'instagram', 'ig': 'instagram', 'yahoo': 'yahoo'}
    med_map = {'google': 'organic', 'line': 'social', 'facebook': 'social',
               'email': 'email', 'instagram': 'social', 'yahoo': 'organic'}

    for row in range(2, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if not a_val or not str(a_val).strip():
            continue

        for col, name in mark_cols.items():
            val = ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == '':
                ws.cell(row=row, column=col).value = '—'

        # 從 Referer 推斷 Campaign
        ref = clean(ws.cell(row=row, column=45).value)
        if ref and ref != '—':
            ref_lower = ref.lower()
            for keyword, source in src_map.items():
                if keyword in ref_lower:
                    cs = ws.cell(row=row, column=46).value
                    if cs == '—':
                        ws.cell(row=row, column=46).value = source
                    cm = ws.cell(row=row, column=47).value
                    if source in med_map and cm == '—':
                        ws.cell(row=row, column=47).value = med_map[source]
                    break

    return True


# ──────────────────────────────────────────────
# Phase 2：轉換為蝦皮格式
# ──────────────────────────────────────────────

def convert_to_shopee(ws_src, ws_tpl):
    """將清理後的官網訂單對應至蝦皮格式工作表"""

    # 清除範本既有資料（保留標題列）
    if ws_tpl.max_row > 1:
        for r in range(2, ws_tpl.max_row + 1):
            for c in range(1, min(ws_tpl.max_column + 1, 56)):
                ws_tpl.cell(row=r, column=c).value = None

    out_row = 2
    total = 0
    for src_row in range(2, ws_src.max_row + 1):
        oid = clean(ws_src.cell(row=src_row, column=1).value)   # 訂單編號
        pn  = clean(ws_src.cell(row=src_row, column=31).value)  # 商品名稱
        if not oid or not pn:
            continue

        addr = clean(ws_src.cell(row=src_row, column=18).value)
        store_code, store_name = parse_store(addr)
        city, area = parse_city_area(addr)

        # 商品編號 & 商品型號（用於主商品貨號 & 商品選項貨號）
        sku   = clean(ws_src.cell(row=src_row, column=32).value)
        model = clean(ws_src.cell(row=src_row, column=35).value)

        row_data = {
            1:  oid,                                          # 訂單編號
            2:  clean(ws_src.cell(row=src_row, column=6).value) or '—',  # 訂單狀態
            3:  '',                                            # 熱門商品
            4:  '',                                            # 退貨/退款狀態
            5:  clean(ws_src.cell(row=src_row, column=12).value) or clean(ws_src.cell(row=src_row, column=10).value),  # 買家帳號
            6:  clean(ws_src.cell(row=src_row, column=3).value)[:10],  # 訂單成立日期
            7:  clean(ws_src.cell(row=src_row, column=25).value),       # 商品總價
            8:  clean(ws_src.cell(row=src_row, column=26).value) or '0',# 買家支付運費
            9:  '',                                            # 蝦皮補助運費
            10: '',                                            # 退貨運費
            11: clean(ws_src.cell(row=src_row, column=30).value),       # 買家總支付金額
            12: '',                                            # 蝦皮補貼金額
            13: '',                                            # 蝦幣折抵
            14: '',                                            # 銀行信用卡活動折抵
            15: clean(ws_src.cell(row=src_row, column=44).value),       # 優惠代碼
            16: '',                                            # 賣場優惠券
            17: '',                                            # 賣家蝦幣回饋券
            18: '',                                            # 優惠券
            19: clean(ws_src.cell(row=src_row, column=28).value),       # 成交手續費
            20: '',                                            # 其他服務費
            21: '',                                            # 金流與系統處理費
            22: '',                                            # 分期付款期數
            23: '',                                            # 金流與系統處理費率
            24: '',                                            # 成交手續費規則名稱
            25: pn,                                            # 商品名稱
            26: sku,                                           # 商品ID
            27: model,                                         # 商品選項名稱
            28: '',                                            # 規格ID
            29: '',                                            # 蝦皮商品編碼
            30: clean(ws_src.cell(row=src_row, column=37).value),       # 商品原價
            31: clean(ws_src.cell(row=src_row, column=37).value),       # 商品活動價格
            32: sku,                                           # 主商品貨號 ← 商品編號
            33: model if model and model != '—' else '—',      # 商品選項貨號 ← 商品型號
            34: clean(ws_src.cell(row=src_row, column=38).value) or '1',# 數量
            35: '',                                            # 退貨數量
            36: '',                                            # 促銷組合指標
            37: '',                                            # 蝦皮促銷組合折扣
            38: addr,                                          # 收件地址
            39: clean(ws_src.cell(row=src_row, column=17).value),       # 收件者電話
            40: '',                                            # 蝦皮專線和包裹查詢碼
            41: store_code,                                    # 取件門市店號
            42: city,                                          # 城市
            43: area,                                          # 行政區
            44: '',                                            # 郵遞區號
            45: clean(ws_src.cell(row=src_row, column=16).value),       # 收件者姓名
            46: map_shipping(clean(ws_src.cell(row=src_row, column=15).value)),  # 寄送方式
            47: '',                                            # 出貨方式
            48: '',                                            # 備貨時間
            49: map_payment(clean(ws_src.cell(row=src_row, column=7).value)),   # 付款方式
            50: '',                                            # 最晚出貨日期
            51: clean(ws_src.cell(row=src_row, column=19).value),       # 包裹查詢號碼
            52: '',                                            # 買家付款時間
            53: clean(ws_src.cell(row=src_row, column=20).value),       # 實際出貨時間
            54: '',                                            # 訂單完成時間
            55: clean(ws_src.cell(row=src_row, column=13).value),       # 買家備註
        }

        for col, val in row_data.items():
            ws_tpl.cell(row=out_row, column=col).value = val
        out_row += 1
        total += 1

    return total


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    # Fix stdout encoding for Windows
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) < 3:
        print("用法: python sc_to_shopee.py <來源訂單.xlsx> <蝦皮格式.xlsx> [輸出檔名.xlsx]")
        print()
        print("範例:")
        print("  python sc_to_shopee.py 訂單.xlsx 蝦皮格式.xlsx")
        print("  python sc_to_shopee.py 訂單.xlsx 蝦皮格式.xlsx 蝦皮格式_訂單_自訂.xlsx")
        sys.exit(1)

    src_file = sys.argv[1]
    tpl_file = sys.argv[2]
    out_file = sys.argv[3] if len(sys.argv) > 3 else (
        os.path.join(os.path.dirname(src_file),
                     f"蝦皮格式_{os.path.splitext(os.path.basename(src_file))[0]}.xlsx")
    )

    if not os.path.exists(src_file):
        print(f"錯誤: 找不到來源檔案: {src_file}")
        sys.exit(1)
    if not os.path.exists(tpl_file):
        print(f"錯誤: 找不到蝦皮格式範本: {tpl_file}")
        sys.exit(1)

    # ── 備份來源 ──
    bak_file = src_file.replace('.xlsx', '_backup.xlsx')
    if not os.path.exists(bak_file):
        shutil.copy2(src_file, bak_file)
        print(f"備份來源 -> {bak_file}")

    print("=" * 50)
    print("[Phase 1] 訂單清理中...")
    wb_src = openpyxl.load_workbook(src_file)
    clean_orders(wb_src.active)
    wb_src.save(src_file)

    print()
    print("[Phase 2] 轉換至蝦皮格式中...")
    wb_src2 = openpyxl.load_workbook(src_file)
    wb_tpl  = openpyxl.load_workbook(tpl_file)

    sheet_name = '工作表1' if '工作表1' in wb_tpl.sheetnames else wb_tpl.sheetnames[0]
    ws_tpl = wb_tpl[sheet_name]

    # 檢查並取消合併儲存格
    merged = list(ws_tpl.merged_cells.ranges)
    for mr in merged:
        ws_tpl.unmerge_cells(str(mr))
    if merged:
        print(f"  -> 已取消 {len(merged)} 個合併儲存格")

    count = convert_to_shopee(wb_src2.active, ws_tpl)
    wb_tpl.save(out_file)

    print()
    print("=" * 50)
    print(f"轉換完成!")
    print(f"  -> 輸出: {out_file}")
    print(f"  -> 筆數: {count} 筆資料，55 個欄位")


if __name__ == '__main__':
    main()
